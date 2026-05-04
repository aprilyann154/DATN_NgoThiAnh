import openpyxl
from openpyxl.styles import Font
import xml.etree.ElementTree as ET
import os

print(f"{'='*50}\nBẮT ĐẦU TRÍCH XUẤT KẾT QUẢ TỪ PYTEST SANG EXCEL\n{'='*50}")

xml_file = "result.xml"
if not os.path.exists(xml_file):
    print(f"[!] LỖI: Không tìm thấy file '{xml_file}'.")
    print("Vui lòng chạy lệnh: python -m pytest tests/ -v --junitxml=result.xml")
    exit()

tree = ET.parse(xml_file)
root = tree.getroot()

test_results = {}
for testcase in root.iter('testcase'):
    test_name = testcase.get('name') # VD: test_TC_QLS_001_open_add_book_form
    status = "Passed"
    error_msg = ""
    
    failure = testcase.find('failure')
    skipped = testcase.find('skipped')
    
    if failure is not None:
        status = "Failed"
        error_msg = failure.get('message', 'Lỗi').split('\n')[0][:100] # Lấy dòng lỗi ngắn gọn
    elif skipped is not None:
        status = "Skipped"
        error_msg = skipped.get('message', 'Bỏ qua').split('\n')[0][:100]
        
    test_results[test_name] = {"status": status, "note": error_msg}

print(f"[*] Đã đọc thành công {len(test_results)} kết quả test từ Pytest.\n")

# ĐIỀN KẾT QUẢ VÀO FILE EXCEL VÀ LƯU THÀNH FILE MỚI

input_excel = "Test_case_DATN.xlsx"
output_excel = "Test_case_DATN_KetQua.xlsx"

try:
    wb = openpyxl.load_workbook(input_excel)
except FileNotFoundError:
    print(f"[!] LỖI: Không tìm thấy file gốc '{input_excel}'")
    exit()

HEADER_ALIASES = {
    "tc_id": ["test case id", "tc id", "id", "mã tc"],
    "result": ["test results", "test result", "kết quả thực tế", "test results edge"], 
    "notes": ["notes", "ghi chú", "bugid"]
}

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    HEADER_ROW = None
    header_map = {}

    for row in range(1, 15):
        temp_map = {}
        for col in range(1, sheet.max_column + 1):
            cell_val = str(sheet.cell(row=row, column=col).value or "").strip().lower()
            if not cell_val: continue
            for key, aliases in HEADER_ALIASES.items():
                if key not in temp_map and any(alias in cell_val for alias in aliases):
                    temp_map[key] = col
        if "tc_id" in temp_map and "result" in temp_map:
            HEADER_ROW = row
            header_map = temp_map
            break

    if not HEADER_ROW: 
        continue 

    print(f"[*] Đang điền kết quả cho Sheet: '{sheet.title}'...")
    filled_count = 0

    for row in range(HEADER_ROW + 1, sheet.max_row + 1):
        tc_id = str(sheet.cell(row=row, column=header_map["tc_id"]).value or "").strip()
        
        if not tc_id or tc_id == "None": continue

        matched_test = None
        for t_name in test_results.keys():
            if tc_id in t_name:
                matched_test = t_name
                break
        
        if matched_test:
            status = test_results[matched_test]["status"]
            note = test_results[matched_test]["note"]
            
            res_cell = sheet.cell(row=row, column=header_map["result"])
            res_cell.value = status
            
            if status == "Passed":
                text_color = "00B050" # Xanh lá cây
            elif status == "Failed":
                text_color = "FF0000" # Đỏ
            elif status == "Skipped":
                text_color = "FFC000" # Vàng cam (Giúp nổi bật hơn màu vàng tươi trên nền trắng)
            else:
                text_color = "000000" # Đen mặc định
                
            res_cell.font = Font(color=text_color, bold=True)
            
            if "notes" in header_map:
                note_cell = sheet.cell(row=row, column=header_map["notes"])
                note_cell.value = note if status in ["Failed", "Skipped"] else ""
            
            filled_count += 1

    print(f"    -> Đã điền {filled_count} cases.")

wb.save(output_excel)
print(f"\n{'='*50}\n[V] HOÀN THÀNH XUẤT SẮC! \nFile báo cáo đã được lưu tại: {output_excel}\n{'='*50}")